#!/usr/bin/env python

import MySQLdb as mdb
import sys
import json

import pprint
import uuid

import time
from datetime import datetime

from openpyxl import load_workbook

pp = pprint.PrettyPrinter(indent = 1)

# Database connectivity
dbname = 'emed'
crd_fname = "EVTM.crd"

# Get credentials
try:
	dbcreds = json.loads(open(crd_fname).read())
except IOError as e:
	print "Error %d: Database Credentials file %s not found" % (e.args[0], crd_fname)
	sys.exit(1)

# the database must exist in the crd file
try: 
	uname = dbcreds[dbname]['uname']
	pword = dbcreds[dbname]['pword']
	try:
		dbhost = dbcreds[dbname]['dbhost']
	except KeyError, e:
		dbhost = 'localhost'
except KeyError, e:
	print "Database named %s not found in %s. No credentials available." % (e, crd_fname)
	sys.exit(1)

# can we connect?
try:
	dbh = mdb.connect(host = dbhost, user = uname,  passwd = pword, db = dbname)
except mdb.Error, e:
	print "Error %d: %s" % (e.args[0], e.args[1])
	sys.exit(1)

# Get a curstor into the database
cur = dbh.cursor(mdb.cursors.DictCursor)

wb = load_workbook(filename = 'AllEvents_20170309-01.xlsx')

# f_schema=open('./emed_schema.sql', "w")
# f_schema.write('USE mysql;\n\n')
# f_schema.write('DROP DATABASE IF EXISTS emed;\n\n')
# f_schema.write('CREATE DATABASE IF NOT EXISTS emed;\n\n')
# f_schema.write('GRANT ALL PRIVILEGES ON emed.* TO \'eventmgt\'@\'%\';\n\n')
# f_schema.write('USE emed;\n\n')

input_sheetname = 'events'
ws = wb[input_sheetname]

skip_header = False
if ws["A1"].value == "AppName":
    skip_header = True

sql_preamble = 'INSERT INTO `events` ( AppName, EventName, AlertMessage, \
ThresholdValue, ThresholdUnit, EventSeverity, EventID, EventMessage, \
AlertName, ThresholdName, EventGUID, AppGUID, AlertGUID, ThresholdGUID \
) VALUES ('
sql_epilogue = ' );'

	
rowcount = 1
for row in ws.rows:
	sql_values = ''
	#print "Row %d" % (rowcount)
	for cell in row:
		if skip_header == True and rowcount != 1:
			#print sql_values
			#sql_values = sql_values + ' \'' + cell.value + '\','
			sql_values = '%s \'%s\',' % (sql_values, cell.value)
			#print(cell.value)
	rowcount = rowcount + 1
	sql_values = sql_values[:-1]
	if skip_header == False or rowcount >= 3:
		sql_statement = sql_preamble + sql_values + sql_epilogue
		#print sql_statement	
		try:
			cur.execute(sql_statement)
		except e:
			print "Error %d: %s" % (e.args[0], e.args[1])
			dbh.rollback

dbh.commit()
#f_schema.close()
