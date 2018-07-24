#!/usr/bin/env python

import MySQLdb as mdb
import traceback
import sys
import time

from openpyxl import Workbook
from openpyxl import load_workbook

import pprint

from emedUtil import emed_ConnectToEMED

pp = pprint.PrettyPrinter(indent = 1, depth = 4)

def loadWStoEMED(type, wsdata, dbh):

	# Get a curstor into the database
	cur = dbh.cursor(mdb.cursors.DictCursor)
	
	wb = load_workbook(filename = wsdata[type]['fname'])
	ws = wb[wsdata[type]['input_sheetname']]
	
	skip_header = False
	if type == 'App' and ws["A1"].value == "AppName":
		skip_header = True
	if type == 'Trap' and ws["A1"].value == "EventName":
		skip_header = True
	if type == 'Internal' and ws["A1"].value == "EventName":
		skip_header = True
	
	if type == 'App':
		sql_preamble = 'INSERT INTO `eventsDynamic` ( AppName, EventName, AlertMessage, \
			ThresholdValue, ThresholdUnit, EventSeverity, EventID, EventMessage, \
			AlertName, ThresholdName, EventGUID, AppGUID, AlertGUID, ThresholdGUID, \
			ThresholdLocalID, Formula \
			) VALUES ('
	
	if type == 'Trap':
		sql_preamble = 'INSERT INTO `eventsTrap` ( 	`EventName`, `EventSeverity`, \
			`EventMessage`,	`EventGUID`, `PowerpackGUID` \
			) VALUES ('
	
	if type == 'Internal':
		sql_preamble = 'INSERT INTO `eventsInternal` ( 	`EventName`, `EventSeverity`, \
			`EventMessage`,	`EventGUID`, `PowerpackGUID` \
			) VALUES ('

	sql_epilogue = ' );'
		
	rowcount = 1
	for row in ws.rows:
		sql_values = ''
		#print "Row %d" % (rowcount)
		for cell in row:
			if skip_header == True and rowcount != 1:
				#pp.pprint(cell.value)
				#pp.pprint(sql_values)
				if isinstance(cell.value, basestring):
					newvalue = cell.value.replace("'", "")
				else:
					newvalue = cell.value
				#sql_values = sql_values + ' \'' + cell.value + '\','
				sql_values = '%s \'%s\',' % (sql_values, newvalue)
				#print(cell.value)
		rowcount = rowcount + 1
		sql_values = sql_values[:-1]
		if skip_header == False or rowcount >= 3:
			sql_statement = sql_preamble + sql_values + sql_epilogue
			#print sql_statement	
			try:
				cur.execute(sql_statement)
			except e:
				print "Error %d: %s" % (e.args[0], e.args[1])
				dbh.rollback
	
	dbh.commit()
	cur.close()
	
def loadEM7DataToEMED(eventType, ddict):

	ldbh = emed_ConnectToEMED('EVTM.crd')

	if eventType == 'Dynamic':
		validTypeFound = True
	
	if eventType == 'Trap':
		validTypeFound = True
	
	if eventType == 'Internal':
		validTypeFound = True
	
	try:
		validTypeFound
	except NameError:
		print "eventType: %s is not defined in function constructSQLForInsertIntoEMED.  Exiting." % (eventType)
		sys.exit(1)

	# We need to remove any existing entries already read in from 
	# EM7 on previous runs to prevent doubles in the output
	cur = ldbh.cursor(mdb.cursors.DictCursor)
	sql_delete = 'DELETE FROM `events%s`;' % (eventType)
	#print sql_delete
	try:
		cur.execute(sql_delete)
	except mdb.Error, e:
		print "Error %d: %s" % (e.args[0], e.args[1])
		print ""
		print sql_delete
		ldbh.rollback
		sys.exit(1)
	
	ldbh.commit()
	cur.close()
	cur = ldbh.cursor(mdb.cursors.DictCursor)
	
	for row in ddict[eventType]['results']:
		#pp.pprint(row)
		sql_insert = constructSQLForInsertIntoEMED(eventType, row)
		#print sql_insert
		try:
			# The library does not return warnings as an exception that can be caught
			# so the try here is not valid.  Leaving it since it does no harm (KCC: 2018-07-24)

			# It also appears that the sql library will cache a given response and only
			# print it once per execution of the db session (or cursor).
			# This might be related to SET WARNING LIMIT in mysql.
			cur.execute(sql_insert)

		except mdb.Error, e:
			print "mdb.Error"
			print "Error %d: %s" % (e.args[0], e.args[1])
			print ""
			print sql_insert
			ldbh.rollback
			sys.exit(1)
		except Exception, e:
			print "Exception"
			print sql_insert
			traceback.print_exc()
			ldbh.rollback
			sys.exit(1)
		
	ldbh.commit()
	cur.close()
	ldbh.close()

def constructSQLForInsertIntoEMED(eventType, d):
	tableName = "events%s" % (eventType)
	keys = ''
	values = ''
	
	# pp.pprint(d)
	
	# The ThresholdValue is a decimal value but may be undefined in EM7 and come over as 'None'
	# Verify that it should be 0.00 by checking ThresholdGUID and ThresholdName
	#print d['ThresholdValue']
#	try:
#		if d['ThresholdGUID'] == None and d['ThresholdName'] == None:
#			d['ThresholdValue'] = 0.00
#	except KeyError:
#		pass

	
	firstIteration = True
	for k in d:

		if isinstance(d[k], str):
			try:
				v = d[k].replace('\r','') # remove CR
	#			v = v.replace('\'','\\\'') # replace single quote with escaped single quote
				v = v.replace('\'','') # remove single quote 
				
				# Note: 2018-Jul-20
				# EM7 has embedded unicode characters in the event message that cause issues with 
				# the mysql library.
				#
				# The following decode/replace recodes the unicode as ASCII  and then looks for 
				# problematic strings that have been observed in the data set being pulled from EM7
				# 
				# This may have to be rethought in the future to find a better solution
				v = v.decode('unicode_escape').encode('ascii','replace')
				v = v.replace('%T?C', '%T deg. C')
				v = v.replace('%T?C ', '%T deg. C ')
				v = v.replace('%V?C', '%V deg. C')
				v = v.replace('%V?C ', '%V deg. C ')
				# End of Note:
			
			except UnicodeDecodeError as ex:
				print "Unicode character found on input string: '%s'" % (v)
				print ex
				traceback.print_exc()
				sys.exit(1)

		else:
			v = d[k]
			
		#if k == 'ThresholdValue' and d[k] == None:
		#	print d[k]
		
		if firstIteration:
			keys = "(`%s`" % (k)
			if v == None:
				values = "(NULL"
			else:
				values = "('%s'" % (v)
			firstIteration = False
		else:
			keys = "%s, `%s`" % (keys, k)
			if v == None:
				fstring = "%s, %s"
				cval = 'NULL'
			else:
				fstring = "%s, '%s'"
				cval = v
			values = fstring % (values, cval)

	keys = "%s)" % (keys)
	values = "%s)" % (values)

	sql = "INSERT INTO `%s` %s VALUES %s" % (tableName, keys, values)

	#print sql
	return sql

def readEventDataFromEM7(eventType, eventTypeData ,dbh):
	try:
		eventTypeData[eventType]['results']
	except KeyError:
		eventTypeData[eventType]['results'] = {}

	#print eventType
	try:
		events_cur = dbh.cursor(mdb.cursors.DictCursor)
		events_cur.execute(eventTypeData[eventType]['query'])
	except mdb.Error, e:
		print "Error %d: %s" % (e.args[0], e.args[1])
		sys.exit(1)

	try:
		eventTypeData[eventType]['results'] = events_cur.fetchall()
	except mdb.Error, e:
		print "Error %d: %s" % (e.args[0], e.args[1])
		sys.exit(1)
	except KeyError, e:
		print "Error: Key %s not found" % (e)
		sys.exit(1)

	#pp.pprint(eventTypeData[eventType]['results'])

def defineEM7ReadQueries(eventTypeList, eventTypeData):
	for eventType in eventTypeList:
		try:
			eventTypeData[eventType]
		except KeyError:
			eventTypeData[eventType] = {}
			eventTypeData[eventType]['query'] = ""

	# Read the event policies related to dynamic apps directly from the database

	eventTypeData['Dynamic']['query'] = "\
	select ev.ename as EventName, da.name as AppName, aa.message as AlertMessage \
	,ath.t_value as ThresholdValue, ath.t_unit as ThresholdUnit \
	,ev.eseverity as EventSeverity \
	,ev.id as EventID, ev.emessage as EventMessage \
	,aa.name as AlertName, ath.name as ThresholdName \
	,ev.event_guid as EventGUID, ev.app_guid as AppGUID \
	,aa.alert_guid as AlertGUID, ath.thresh_guid as ThresholdGUID \
	from policies_events as ev \
	inner join dynamic_app as da on ev.app_guid = da.app_guid \
	inner join dynamic_app_alerts as aa on ev.Xoid = aa.alert_id \
	left outer join dynamic_app_thresholds as ath on aa.app_guid = ath.app_guid \
	order by AppName, EventGUID \
	"

	eventTypeData['Trap']['query'] = "\
	select ev.ename as EventName \
	,ev.eseverity as EventSeverity \
	,ev.emessage as EventMessage \
	,ev.event_guid as EventGUID \
	,ev.ppguid as PowerPackGUID \
	from policies_events as ev \
	where ev.esource = (select esource from master.definitions_event_sources where descr = 'Trap') \
	and ev.ppguid is not null \
	order by EventName, EventSeverity \
	"

	eventTypeData['Internal']['query'] = "\
	select policies_events.ename as EventName \
	,policies_events.eseverity as EventSeverity \
	,definitions_internal_messages.msg_txt as EventMessage \
	,policies_events.event_guid as EventGUID \
	,policies_events.ppguid as PowerpackGUID \
	from policies_events \
	RIGHT OUTER JOIN definitions_internal_messages on msg_id = Xoid \
	where policies_events.esource = (select esource from master.definitions_event_sources where descr = 'Internal') \
	and policies_events.eseverity != 0 \
	ORDER BY ename \
	"

	for eventType in eventTypeData:
		if len(eventTypeData[eventType]['query']) == 0:
			raise ValueError, eventType
			