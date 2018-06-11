#!/usr/bin/env python

import MySQLdb as mdb
import sys
import traceback
import json
import re

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl import load_workbook

import pprint

def emed_ConnectToSQL(dbname, crd_fname):
	# Database connectivity
	#dbname = 'emed'
	#crd_fname = "EVTM.crd"

	# Get credentials
	try:
		dbcreds = json.loads(open(crd_fname).read())
	except IOError as e:
		print "Error %d: Database Credentials file %s not found" % (e.args[0], crd_fname)
		sys.exit(1)

	# the database must exist in the crd file
	try: 
		uname = dbcreds[dbname]['uname']
		pword = dbcreds[dbname]['pword']
		try:
			dbhost = dbcreds[dbname]['dbhost']
		except KeyError, e:
			dbhost = 'localhost'
	except KeyError, e:
		print "Database named %s not found in %s. No credentials available." % (e, crd_fname)
		sys.exit(1)

	# can we connect?
	try:
		dbh = mdb.connect(host = dbhost, user = uname,  passwd = pword, db = dbname)
	except mdb.Error, e:
		print "Error %d: %s" % (e.args[0], e.args[1])
		sys.exit(1)

	return dbh
	


#eventtypes = ('eventsApp', 'eventsTrap', 'eventsTrapVarbind', 'eventsInternal')
#eventtypes = emed_getEventTypesFromSQL(dbh)
dbh = emed_ConnectToSQL('emed', 'EVTM.crd')

pp = pprint.PrettyPrinter(indent = 1, depth = 6)

def createDocSet(doctypes, docMetaData, sheets):
	#pp.pprint(docMetaData)
	for doctype in doctypes:
		metaData = docMetaData[doctype]
		#pp.pprint(metaData)
		
		if doctype == 'itsmpriority':
			createITSMPriorityDoc(sheets, docMetaData, metaData)
		else:
			wb = Workbook()
	
			createTitleWS(wb, docMetaData, metaData)
			for sheet in sheets:
				if doctype == 'baseline':
					createBaselineWS(wb, sheet, paletteName=metaData['docControl']['docPalette'], formatWB=metaData['isFormatted'])
				elif doctype == 'procedure':
					createProcedureWS(wb, sheet, paletteName=metaData['docControl']['docPalette'], formatWB=metaData['isFormatted'])
	
			wb.save((('output/%s') % (metaData['fname'])))
		
		print('Created file: %s' % (metaData['fname']))

def createITSMPriorityDoc(sheets, docMetaData, metaData):
	#lpathname = './%s' % (docMetaData['uuid'])
	#os.makedirs(lpathname)
	xmlfile = file((('output/%s') % (metaData['fname'])), 'w', 0)
	#xmlfile.write(('<ITSMPRIORITY uuid=\'%s\' revisionTime='\%s\' fileNameTime=\'%s\'>' % (docMetaData['uuid'], docMetaData['revisionTime'], docMetaData['fileNameTime'])))
	#xmlfile.write('<ITSMPRIORITY uuid="%s" revisionTime="%s" fileNameTime="%s">\n' % (docMetaData['uuid'], docMetaData['revisionTime'], docMetaData['fileNameTime']))
	xmlfile.write('<ITSMPRIORITY>\n')
	xmlfile.write('  <UUID>%s</UUID>\n' % (docMetaData['uuid']))
	xmlfile.write('  <REVISIONTIME>%s</REVISIONTIME>\n' % (docMetaData['revisionTime']))
	xmlfile.write('  <FILENAMETIME>%s</FILENAMETIME>\n' % (docMetaData['fileNameTime']))
	xmlfile.write('  <LOOKUP_TABLE>\n')
	GUIDAlreadySeen = {}
	for sheet in sheets:
		#pp.pprint(sheet)
		try:
			for EventGUID in sheet['Dynamic']['data']:
				try:
					GUIDAlreadySeen[EventGUID]
				except KeyError:
					GUIDAlreadySeen[EventGUID]={}
					#pp.pprint(sheet['Dynamic']['data'][EventGUID])
				
					# extract just the numeric priority
					m = re.search("\d", sheet['Dynamic']['data'][EventGUID]['incidentPriority'])
					if m:
						incPriority = "%c" % (sheet['Dynamic']['data'][EventGUID]['incidentPriority'][m.start()])
					else:
						incPriority = '4'
					
					xmlfile.write('    <ENTRY>\n')
					xmlfile.write('      <GLOBAL_POLICY_NAME>\n')
					xmlfile.write('      %s\n' % (sheet['Dynamic']['data'][EventGUID]['EventName']))
					xmlfile.write('      </GLOBAL_POLICY_NAME>\n')
					xmlfile.write('      <GLOBAL_POLICY_ID>\n')
					xmlfile.write('      %s\n' % (EventGUID))
					xmlfile.write('      </GLOBAL_POLICY_ID>\n')
					xmlfile.write('      <GLOBAL_POLICY_INCIDENT_PRIORITY>\n')
					xmlfile.write('      %s\n' % (incPriority))
					xmlfile.write('      </GLOBAL_POLICY_INCIDENT_PRIORITY>\n')
					xmlfile.write('    </ENTRY>\n')
		except KeyError:
			pass
	xmlfile.write('  </LOOKUP_TABLE>\n')
	xmlfile.write('  <OVERRIDE_TABLE>\n')
	xmlfile.write('  </OVERRIDE_TABLE>\n')
	xmlfile.write('</ITSMPRIORITY>\n')
	xmlfile.close()
	#createITSMPriorityTitle(xfd, sheets, docMetaData, metaData)
	#pass

def createTitleWS(wb, docMetaData, metaData):

	ws = wb.active
	ws.title = "Title"  # This is the tab name in the workbook

	palette = createPalette(metaData['docControl']['docPalette'])
	
	# Create the title worksheet title row
	columnList = ('A', )
	labelList = ('Virtustream', )
	createWSTitleRow(ws, palette, columnList, labelList, format=metaData['isFormatted'], merge=('A', 'B'))
	
	# Create the title worksheet header row
	columnList = ('A', )
	labelList = (metaData['docControl']['docDescTitle'], )
	createWSHeaderRow(ws, palette, columnList, labelList, format=metaData['isFormatted'], merge=('A', 'B'))

	# Create the title worksheet data rows
	titleSheetRows = (
	('Generator', docMetaData['generator'], ),
	('Creation Time', docMetaData['revisionTime'], ),
	('Identifier', docMetaData['uuid'], ),
	)

	rownumber = palette['titledatarow']
	for titleSheetRow in titleSheetRows:
		columnList = ('A', 'B')
		dataList = (titleSheetRow[0], titleSheetRow[1], )
		rownumber = createWSRow(ws, palette, columnList, dataList, 'titledata',\
			rownumber = rownumber, format=metaData['isFormatted'])

	ws.column_dimensions['A'].width = 16
	ws.column_dimensions['B'].width = 50
	
def createBaselineWS(wb, s, paletteName = 'soft', formatWB=True):
	pp = pprint.PrettyPrinter(indent = 1)
	# pp.pprint(s)

	palette = createPalette(paletteName)
	
	severityMapping = ["Healthy" \
	,'Informational' \
	,'Minor' \
	,'Major' \
	,'Critical']
	
	ws = wb.create_sheet(s['SheetName'])

	# Create the monitored object worksheet title row
	columnList = ('A', 'B', 'F', )
	labelList = ('Virtustream', s['SheetDesc'], '', )
	createWSTitleRow(ws, palette, columnList, labelList, format=formatWB, merge = ('B','F'))
	
	# Create the monitored object worksheet header row
	columnList = ('A', 'B', 'C', 'D', 'E', 'F', )
	labelList = ('Event Name', 'Message Format', 'Threshold', 'Unit', 'Severity', 'Inc. Priority', )
	createWSHeaderRow(ws, palette, columnList, labelList, format=formatWB)
	
	# Create the monitored object worksheet data rows

	rownumber = palette['datarow']
	columnList = ('A', 'B', 'C', 'D', 'E', 'F', )

	# Get the list of events for each sheet and type in the sheet
	eventtypes = emed_getEventTypesFromSQL(dbh)
	#pp.pprint(eventtypes)
	for eventtype in eventtypes:
		try:
			localEventTree = s[eventtype['descr']]['data']
			#pp.pprint(localEventTree)

		except KeyError, e:
			# Not all sheets will have all event types so continue
			continue

		try:
			for eventID in s[eventtype['descr']]['data']:
				if isinstance(s[eventtype['descr']]['data'][eventID]['EventSeverity'], int):
					loopSeverity = severityMapping[s[eventtype['descr']]['data'][eventID]['EventSeverity']]
				else:
					loopSeverity = s[eventtype['descr']]['data'][eventID]['EventSeverity']
				
				incidentPriority = s[eventtype['descr']]['data'][eventID]['incidentPriority']
				
				dataList = (s[eventtype['descr']]['data'][eventID]['EventName'],
				s[eventtype['descr']]['data'][eventID]['AlertMessage'],
				str(s[eventtype['descr']]['data'][eventID]['ThresholdValue']),
				s[eventtype['descr']]['data'][eventID]['ThresholdUnit'],
				loopSeverity,
				incidentPriority, 
				)
				
				#pp.pprint(dataList)
				rownumber = createWSRow(ws, palette, columnList, dataList, 'data',\
					rownumber = rownumber, format=formatWB)
		except KeyError, e:
			# A key error here means we didn't get all the data back from the data structure
			# Print the key and exit
			print "KeyError: Missing field %s" % (str(e))
			traceback.print_stack()
			sys.exit(1)
			
	# Set the appropriate column widths for this sheet
	ws.column_dimensions['A'].width = 55
	ws.column_dimensions['B'].width = 105
	ws.column_dimensions['C'].width = 10
	ws.column_dimensions['D'].width = 10
	ws.column_dimensions['E'].width = 10
	ws.column_dimensions['F'].width = 15

#	filterStart_ID = 'A%d'% (palette['headerrow'])
#	filterEnd_ID = 'F%d' % (row - 1)
#	filterRange = '%s:%s' % (filterStart_ID, filterEnd_ID)
#	ws.auto_filter_ref = filterRange
#	ws.auto_filter.add_filter_column('A')
#	ws.auto_filter.add_filter_column('B')
#	ws.auto_filter.add_filter_column('C')
#	ws.auto_filter.add_filter_column('D')
#	ws.auto_filter.add_filter_column('E')
#	ws.auto_filter.add_filter_column('F')

#	pp.pprint(ws)

def createProcedureWS(wb, s, paletteName = 'soft', formatWB=True):
	pp = pprint.PrettyPrinter(indent = 1)
	# pp.pprint(s)

	palette = createPalette(paletteName)
	
	severityMapping = ["Healthy" \
	,'Informational' \
	,'Minor' \
	,'Major' \
	,'Critical']
	
	ws = wb.create_sheet(s['SheetName'])

	# Create the monitored object worksheet title row
	columnList = ('A', 'B', 'F', )
	labelList = ('Virtustream', s['SheetDesc'], '', )
	createWSTitleRow(ws, palette, columnList, labelList, format=formatWB, merge = ('B','F'))
	
	# Create the monitored object worksheet header row
	columnList = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', )
	labelList = ('Event Name', 'Message Format', 'Threshold', 'Unit', 'Severity', 'Source', 'Inc. Priority', 'Procedure', 'Event GUID')
	createWSHeaderRow(ws, palette, columnList, labelList, format=formatWB)
	
	# Create the monitored object worksheet data rows

	rownumber = palette['datarow']
	columnList = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', )
	
	# Get the list of events for each sheet and type in the sheet
	eventtypes = emed_getEventTypesFromSQL(dbh)
	for eventtype in eventtypes:
		try:
			localEventTree = s[eventtype['descr']]['data']
			#pp.pprint(localEventTree)

		except KeyError, e:
			# Not all sheets will have all event types so continue
			continue

		try:
			for eventID in s[eventtype['descr']]['data']:
				# Normalize loop severity
				if isinstance(s[eventtype['descr']]['data'][eventID]['EventSeverity'], int):
					loopSeverity = severityMapping[s[eventtype['descr']]['data'][eventID]['EventSeverity']]
				else:
					loopSeverity = s[eventtype['descr']]['data'][eventID]['EventSeverity']

				incidentPriority = s[eventtype['descr']]['data'][eventID]['incidentPriority']
				procText = 	s[eventtype['descr']]['data'][eventID]['procText']
				
				dataList = (s[eventtype['descr']]['data'][eventID]['EventName'],
				s[eventtype['descr']]['data'][eventID]['AlertMessage'],
				str(s[eventtype['descr']]['data'][eventID]['ThresholdValue']),
				s[eventtype['descr']]['data'][eventID]['ThresholdUnit'],
				loopSeverity,
				s[eventtype['descr']]['data'][eventID]['AppName'],
				incidentPriority,
				procText, 
				eventID
				)
				
				#pp.pprint(dataList)
				rownumber = createWSRow(ws, palette, columnList, dataList, 'data',\
					rownumber = rownumber, format=formatWB)

		except KeyError, e:
			# A key error here means we didn't get all the data back from the data structure
			# Print the key and exit
			print "KeyError: Missing field %s" % (str(e))
			traceback.print_stack()
			sys.exit(1)

	# Set the appropriate column widths for this sheet
	ws.column_dimensions['A'].width = 55
	ws.column_dimensions['B'].width = 105
	ws.column_dimensions['C'].width = 10
	ws.column_dimensions['D'].width = 10
	ws.column_dimensions['E'].width = 10
	ws.column_dimensions['F'].width = 35
	ws.column_dimensions['G'].width = 15
	ws.column_dimensions['H'].width = 35
	ws.column_dimensions['I'].width = 35

#	filterStart_ID = 'A%d'% (palette['headerrow'])
#	filterEnd_ID = 'F%d' % (row - 1)
#	filterRange = '%s:%s' % (filterStart_ID, filterEnd_ID)
#	ws.auto_filter_ref = filterRange
#	ws.auto_filter.add_filter_column('A')
#	ws.auto_filter.add_filter_column('B')
#	ws.auto_filter.add_filter_column('C')
#	ws.auto_filter.add_filter_column('D')
#	ws.auto_filter.add_filter_column('E')
#	ws.auto_filter.add_filter_column('F')

#	pp.pprint(ws)

def createWSTitleRow(ws, palette, columnList, labelList, merge = None, format=True):
	createWSRow(ws, palette, columnList, labelList, 'title', merge=merge, format=format)

def createWSHeaderRow(ws, palette, columnList, labelList, merge = None, format=True):
	createWSRow(ws, palette, columnList, labelList, 'header', merge=merge, format=format)
	
def createWSRow(ws, palette, columnList, labelList, rowtype, rownumber=None, merge=None, format=True):
	rowname = '%srow' % (rowtype)
	fontsizename = '%sfontsize' % (rowtype)
	fontboldname = '%sfontbold' % (rowtype)
	
	if rowtype != 'data':
		fillcolorname = '%sfillcolor' % (rowtype)
	else:
		fillcolorname = '%sfillcolor%d' % (rowtype, (rownumber % 2))

	bold = palette[fontboldname]
		
	if rownumber:
		rowvalue = str(rownumber)
	else:
		rowvalue = palette[rowname]

	for idx in range(0, len(columnList)):
		currentCellName = '%s%s' % (columnList[idx], rowvalue)
		currentCell = ws[currentCellName]
		currentCell.value = labelList[idx]
		try:
			currentCell.font = Font(bold = bold, size=palette[fontsizename])
		except:
			print('KeyError: %s not found in palette' % (fontsizename))
			
		try:
			if format:
				currentCell.fill = PatternFill('solid', fgColor=palette[fillcolorname])
		except KeyError:
			print('KeyError: %s not found in palette' % (fillcolorname))
			
	#	ws.merge_cells('%s:%s' % (cellB_ID, cellF_ID))
	if merge:
		mergeRange = '%s%d:%s%d' % (merge[0], palette[rowname], merge[1], palette[rowname])
		ws.merge_cells(mergeRange)
		
	if rownumber:
		return int( int(rownumber) + 1)

def createPalette(style):
	palette = {}

	palette['dell'] = {}

	palette['dell']['titlerow'] = 1
	palette['dell']['titlefontsize'] = 14
	palette['dell']['titlefontbold'] = True
	palette['dell']['titlefillcolor'] = 'dcebe1'
	
	palette['dell']['titledatarow'] = 6
	palette['dell']['titledatafontsize'] = 11
	palette['dell']['titledatafontbold'] = False
	palette['dell']['titledatafillcolor'] = 'ffffff'
	
	palette['dell']['headerrow'] = 2
	palette['dell']['headerfontsize'] = 12
	palette['dell']['headerfontbold'] = True
	palette['dell']['headerfillcolor'] = '007db8'

	palette['dell']['datarow'] = 3
	palette['dell']['datafontsize'] = 11
	palette['dell']['datafontbold'] = False

	palette['dell']['datafillcolor0'] = 'ffffff'
	palette['dell']['datafillcolor1'] = 'dcebe1'	

	palette['virtustream'] = {}

	palette['virtustream']['titlerow'] = 1
	palette['virtustream']['titlefontsize'] = 14
	palette['virtustream']['titlefontbold'] = True
	palette['virtustream']['titlefillcolor'] = 'bd1e30'
	
	palette['virtustream']['titledatarow'] = 6
	palette['virtustream']['titledatafontsize'] = 11
	palette['virtustream']['titledatafontbold'] = False
	palette['virtustream']['titledatafillcolor'] = 'ffffff'
	
	palette['virtustream']['headerrow'] = 2
	palette['virtustream']['headerfontsize'] = 12
	palette['virtustream']['headerfontbold'] = True
	palette['virtustream']['headerfillcolor'] = '1a4b5f'

	palette['virtustream']['datarow'] = 3
	palette['virtustream']['datafontsize'] = 11
	palette['virtustream']['datafontbold'] = False

	palette['virtustream']['datafillcolor0'] = '586a6f'
	palette['virtustream']['datafillcolor1'] = 'b5b9bb'
	
	palette['soft'] = {}

	palette['soft']['titlerow'] = 1
	palette['soft']['titlefontsize'] = 14
	palette['soft']['titlefontbold'] = True
	palette['soft']['titlefillcolor'] = 'f6273e'

	palette['soft']['titledatarow'] = 6
	palette['soft']['titledatafontsize'] = 11
	palette['soft']['titledatafontbold'] = False
	palette['soft']['titledatafillcolor'] = 'ffffff'

	palette['soft']['headerrow'] = 2
	palette['soft']['headerfontsize'] = 12
	palette['soft']['headerfontbold'] = True
	palette['soft']['headerfillcolor'] = '21627c'

	palette['soft']['datarow'] = 3
	palette['soft']['datafontsize'] = 11
	palette['soft']['datafontbold'] = False

	palette['soft']['datafillcolor0'] = '728a90'
	palette['soft']['datafillcolor1'] = 'd9d9d9'
	
	return palette[style]
	
def emed_getEventsTrapVarbindTableName(dbh, SheetGUID, current_type):
	tablename_prequery = "select dataIdentifier from sheetMapping\
		where sheetMapping.DataType = %d and sheetMapping.SheetGUID = %s"
	try:
		tablename_cur = dbh.cursor(mdb.cursors.DictCursor)
		tablename_query =  tablename_prequery % ( int(current_type), SheetGUID )
		tablename_cur.execute(tablename_query)
	except mdb.Error, e:
		print "Error %d: %s" % (e.args[0], e.args[1])
		sys.exit(1)

	try:
		tablename = tablename_cur.fetchone()
	except mdb.Error, e:
		print "Error %d: %s" % (e.args[0], e.args[1])
		sys.exit(1)
	
	tablename_cur.close()
	return tablename
	
def	emed_getEventDetailsFromSQL(dbh, event_query):
	#pp.pprint('emed_getEventDetailsFromSQL')
	try:
		event_cur = dbh.cursor(mdb.cursors.DictCursor)
		#pp.pprint(event_query)
		event_cur.execute(event_query)
	except mdb.Error, e:
		print "Error %d: %s" % (e.args[0], e.args[1])
		sys.exit(1)
	
	try:
		event = event_cur.fetchone()
	except mdb.Error, e:
		print "Error %d: %s" % (e.args[0], e.args[1])
		sys.exit(1)
				
	event_cur.close()
	
	#pp.pprint(event)
	return event

	
def	emed_getEventDetails_App(dbh, eventGUID, eventRoot):
	#pp.pprint('emed_getEventDetails_App')
	#event_prequery = "SELECT EventName, AlertMessage, ThresholdValue, ThresholdUnit, EventSeverity, AppName \
	#from eventsApp where EventGUID = '%s'"
	event_prequery = "SELECT EventName, AlertMessage, ThresholdValue, ThresholdUnit, EventSeverity, AppName \
		, incidentPriority, procText.procText \
		from eventsApp \
		LEFT OUTER JOIN eventProcMapping on eventsApp.EventGUID = eventProcMapping.EventGUID \
		LEFT OUTER JOIN procText on procText.procTextID = eventProcMapping.procTextID \
		where eventsApp.EventGUID = '%s'"
	event_query = event_prequery % (eventGUID)
	return emed_getEventDetailsFromSQL(dbh,event_query)

def emed_getEventDetails_Trap(dbh, eventGUID, eventRoot):
	#pp.pprint('emed_getEventDetails_Trap')
	#event_prequery = "SELECT EventName, EventMessage as AlertMessage, EventSeverity \
	#from eventsTrap where EventGUID = '%s'"
	event_prequery = "SELECT EventName, EventMessage as AlertMessage, EventSeverity, incidentPriority, procText.procText \
	from eventsTrap \
	LEFT OUTER JOIN eventProcMapping on eventsTrap.EventGUID = eventProcMapping.EventGUID \
	LEFT OUTER JOIN procText on procText.procTextID = eventProcMapping.procTextID \
	where eventsTrap.EventGUID = '%s'"
	event_query = event_prequery % (eventGUID)
	return emed_getEventDetailsFromSQL(dbh,event_query)

def emed_getEventDetails_TrapVarbind(dbh, eventGUID, eventRoot, tablename):
	#pp.pprint('emed_getEventDetails_TrapVarbind')
	#pp.pprint(tablename)
	#event_prequery = "SELECT trapName as EventName, trapMessage as AlertMessage,\
	#	trapSeverity as EventSeverity from %s where trapCode = '%%s'" % (tablename)
	event_prequery = "SELECT trapName as EventName, trapMessage as AlertMessage, trapSeverity as EventSeverity \
	, incidentPriority, procText from %s \
	LEFT OUTER JOIN eventProcMapping on %s.trapCode = eventProcMapping.EventGUID \
	LEFT OUTER JOIN procText on procText.procTextID = eventProcMapping.procTextID \
	where trapCode = '%%s'" % (tablename, tablename)
	event_query = event_prequery % (eventGUID)
	return emed_getEventDetailsFromSQL(dbh,event_query)

def emed_getEventDetails_Internal(dbh, eventGUID, eventRoot):
	#pp.pprint('emed_getEventDetails_Internal')
	#event_prequery = "SELECT EventName, EventMessage as AlertMessage, EventSeverity \
	#from eventsTrap where EventGUID = '%s'"
	event_prequery = "SELECT EventName, EventMessage as AlertMessage, EventSeverity, incidentPriority, procText.procText \
	from eventsInternal \
	LEFT OUTER JOIN eventProcMapping on eventsInternal.EventGUID = eventProcMapping.EventGUID \
	LEFT OUTER JOIN procText on procText.procTextID = eventProcMapping.procTextID \
	where eventsInternal.EventGUID = '%s'"
	event_query = event_prequery % (eventGUID)
	return emed_getEventDetailsFromSQL(dbh,event_query)

def emed_getEventDetails(dbh, eventtype, eventGUID, eventRoot):
	eventRoot['data'][eventGUID] = {}
	if 'Dynamic' == eventtype:
		eventRoot['data'][eventGUID] = emed_getEventDetails_App(dbh, eventGUID, eventRoot)
	elif 'TrapByOID' == eventtype:
		eventRoot['data'][eventGUID] = emed_getEventDetails_Trap(dbh, eventGUID, eventRoot)
	elif 'TrapByVarbind' == eventtype:
		eventRoot['data'][eventGUID] = emed_getEventDetails_TrapVarbind(dbh, eventGUID, eventRoot, eventRoot['source']['tablename'])
	elif 'Internal' == eventtype:
		eventRoot['data'][eventGUID] = emed_getEventDetails_Internal(dbh, eventGUID, eventRoot)
	elif 'SingleEvents' == eventtype:
		pass
	else:
		print "Unidentified Event Type: %s in function emed_getEventDetails" % (eventtype)
		sys.exit(1)

	if 'Dynamic' == eventtype:
		if eventRoot['data'][eventGUID]['ThresholdUnit'] == 'None' and eventRoot['data'][eventGUID]['AppName'][:4] == 'VNXe':
			eventRoot['data'][eventGUID]['ThresholdUnit'] = ''
			eventRoot['data'][eventGUID]['ThresholdValue'] = ''
	
		if eventRoot['data'][eventGUID]['ThresholdUnit'] == '&#8451':
			eventRoot['data'][eventGUID]['ThresholdUnit'] = 'Celsius'
		if eventRoot['data'][eventGUID]['ThresholdUnit'] == 'NULL':
			eventRoot['data'][eventGUID]['ThresholdUnit'] = ''
		if eventRoot['data'][eventGUID]['ThresholdUnit'] == 'None':
			eventRoot['data'][eventGUID]['ThresholdUnit'] = ''
	
		eventRoot['data'][eventGUID]['AppName'] = "%s (DynApp)" % (eventRoot['data'][eventGUID]['AppName'])
	
	elif 'TrapByOID' == eventtype:
		eventRoot['data'][eventGUID]['ThresholdValue'] = ''
		eventRoot['data'][eventGUID]['ThresholdUnit'] = ''
		eventRoot['data'][eventGUID]['AppName'] = 'SNMP Trap by OID'
	
	elif 'TrapByVarbind' == eventtype:
		eventRoot['data'][eventGUID]['ThresholdValue'] = ''
		eventRoot['data'][eventGUID]['ThresholdUnit'] = ''
		eventRoot['data'][eventGUID]['AppName'] = 'SNMP Trap by Varbind'
	
	elif 'Internal' == eventtype:
		eventRoot['data'][eventGUID]['ThresholdValue'] = ''
		eventRoot['data'][eventGUID]['ThresholdUnit'] = ''
		eventRoot['data'][eventGUID]['AppName'] = 'Internal'
	
	elif 'SingleEvents' == eventtype:
		pass
	else:
		raise TypeError
	
def emed_getEventsDetails(dbh, eventtype, eventGUID, eventRoot):
	emed_getEventDetails(dbh, eventtype, eventGUID, eventRoot)

def emed_getEventTypesFromSQL(dbh):
	eventtypes_query = "select esource, descr, eventTypeActive from definitions_event_sources where eventTypeActive = TRUE order by esource"
	eventtypes_cur = dbh.cursor(mdb.cursors.DictCursor)
	try:
		eventtypes_cur.execute(eventtypes_query)
	except mdb.Error, e:
		print "Error %d: %s" % (e.args[0], e.args[1])
		sys.exit(1)
	try:
		eventtypes = eventtypes_cur.fetchall()
	except mdb.Error, e:
		print "Error %d: %s" % (e.args[0], e.args[1])
		sys.exit(1)
	
	return eventtypes

