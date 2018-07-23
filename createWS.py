#!/usr/bin/env python

import MySQLdb as mdb
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
import sys

import time
from datetime import datetime
import uuid

import pprint

def createTitleWS(wb, revtime, uuidstr, FormatWB=True, generator='emed-blgen.py'):
	ws = wb.active
	ws.title = "Title"

	titlefontsize = 14
	titlefillcolor = 'bd1e30'

	labelfontsize = 12
	labelfillcolor = '1a4b5f'
	
	c=ws['A1']
	c.value = 'Virtustream'
	if FormatWB:
		c.font = Font(bold = True, size=titlefontsize)
		c.fill = PatternFill('solid', fgColor=titlefillcolor)
	ws.merge_cells('A1:B1')
	
	c=ws['A2']
	c.value = 'Monitored Objects and Thresholds for Vblock Systems'
	if FormatWB:
		c.font = Font(bold = True, size=labelfontsize)
		c.fill = PatternFill('solid', fgColor=labelfillcolor)
	ws.merge_cells('A2:B2')
	
	c=ws['A6']
	c.value = 'Generator'
	if FormatWB:
		c.font = Font(size=labelfontsize)
	c=ws['B6']
	c.value = generator
	c=ws['A7']
	c.value = 'Creation Time'
	if FormatWB:
		c.font = Font(size=labelfontsize)
	c=ws['B7'] = revtime
	c=ws['A8']
	c.value = 'Identifier'
	if FormatWB:
		c.font = Font(size=labelfontsize)
	c=ws['B8']
	c.value = uuidstr
	
	
	
def createWS(wb, s, dbh, FormatWB=True):
	pp = pprint.PrettyPrinter(indent = 1)
	# pp.pprint(s)

	# Create styles and positions in worksheets
	titlerow = 1
	#vs_title = NamedStyle(name='VS_Title')
	titlefontsize = 14
	#vs_title.font = Font(bold=True, size=titlefontsize)
	titlefillcolor = 'bd1e30'
	#vs_title.fill = PatternFill("solid", fgColor=titlefillcolor)
	#wb.add_named_style(vs_title)

	headerrow = 2
	headerfontsize = 12
	headerfillcolor = '1a4b5f'

	firstdatarow = 3

	data0fillcolor = '586a6f'
	data1fillcolor = 'b5b9bb'
	
	# create styles
	#style_header = NamedStyle(name='VS_Header')
	#style_data1 = NamedStyle(name='VS_Data0')
	#style_data2 = NamedStyle(name='VS_Data1')
	
	
	
	severityMapping = ["Healthy" \
	,'Informational' \
	,'Minor' \
	,'Major' \
	,'Critical']
	
	ws = wb.create_sheet(s['SheetName'])

	# Title Row
	
	cellA_ID = 'A%d'% (titlerow)
	cellB_ID = 'B%d'% (titlerow)
	cellF_ID = 'F%d'% (titlerow)
	c = ws[cellA_ID]
	c.value = 'Virtustream'
	if FormatWB:
		c.font = Font(bold = True, size=titlefontsize)
		c.fill = PatternFill('solid', fgColor=titlefillcolor)
#	c.style = vs_title


	c = ws[cellB_ID]
	c.value = s['SheetDesc']
	if FormatWB:
		c.font = Font(bold = True, size=titlefontsize)
		c.fill = PatternFill('solid', fgColor=titlefillcolor)
	ws.merge_cells('%s:%s' % (cellB_ID, cellF_ID))
	
	
	# Header Row
	cellA_ID = 'A%d'% (headerrow)
	cellB_ID = 'B%d'% (headerrow)
	cellC_ID = 'C%d'% (headerrow)
	cellD_ID = 'D%d'% (headerrow)
	cellE_ID = 'E%d'% (headerrow)
	cellF_ID = 'F%d'% (headerrow)
	ws[cellA_ID] = 'Event Name'
	c = ws[cellA_ID]
	if FormatWB:
		c.font = Font(bold = True, size=headerfontsize)
		c.fill = PatternFill('solid', fgColor=headerfillcolor)
	ws[cellB_ID] = 'Message Format'
	c = ws[cellB_ID]
	if FormatWB:
		c.font = Font(bold = True, size=headerfontsize)
		c.fill = PatternFill('solid', fgColor=headerfillcolor)
	ws[cellC_ID] = 'Threshold'
	c = ws[cellC_ID]
	if FormatWB:
		c.font = Font(bold = True, size=headerfontsize)
		c.fill = PatternFill('solid', fgColor=headerfillcolor)
	ws[cellD_ID] = 'Unit'
	c = ws[cellD_ID]
	if FormatWB:
		c.font = Font(bold = True, size=headerfontsize)
		c.fill = PatternFill('solid', fgColor=headerfillcolor)
	ws[cellE_ID] = 'Severity'
	c = ws[cellE_ID]
	if FormatWB:
		c.font = Font(bold = True, size=headerfontsize)
		c.fill = PatternFill('solid', fgColor=headerfillcolor)
	ws[cellF_ID] = 'Dynamic App'
	c = ws[cellF_ID]
	if FormatWB:
		c.font = Font(bold = True, size=headerfontsize)
		c.fill = PatternFill('solid', fgColor=headerfillcolor)

	row = firstdatarow
	for eventID in s['eventsDynamic']:
		cellA_ID = 'A%d'% (row)
		cellB_ID = 'B%d'% (row)
		cellC_ID = 'C%d'% (row)
		cellD_ID = 'D%d'% (row)
		cellE_ID = 'E%d'% (row)
		cellF_ID = 'F%d'% (row)
		
		if (row % 2) == 1:
			rowfillcolor = data0fillcolor
		else:
			rowfillcolor = data1fillcolor

		# Get the data for the row from the eventsDynamic table
		try:
			event_cur = dbh.cursor(mdb.cursors.DictCursor)
			event_query = "SELECT \
			EventName, AlertMessage, ThresholdValue, ThresholdUnit, EventSeverity, AppName \
			from eventsDynamic where EventGUID = '%s'" % (eventID)
			event_cur.execute(event_query)
		except mdb.Error, e:
			print "Error %d: %s" % (e.args[0], e.args[1])
			sys.exit(1)

		try:
			event = event_cur.fetchone()
		except mdb.Error, e:
			print "Error %d: %s" % (e.args[0], e.args[1])
			sys.exit(1)

		ws[cellA_ID] = event['EventName']
		c=ws[cellA_ID]
		if FormatWB:
			c.fill = PatternFill('solid', fgColor=rowfillcolor)
		ws[cellB_ID] = event['AlertMessage']
		c=ws[cellB_ID]
		if FormatWB:
			c.fill = PatternFill('solid', fgColor=rowfillcolor)
		ws[cellC_ID] = event['ThresholdValue']
		c=ws[cellC_ID]
		if FormatWB:
			c.fill = PatternFill('solid', fgColor=rowfillcolor)
		if event['ThresholdUnit'] != 'NULL':
			ws[cellD_ID] = event['ThresholdUnit']
		c=ws[cellD_ID]
		if FormatWB:
			c.fill = PatternFill('solid', fgColor=rowfillcolor)
		ws[cellE_ID] = severityMapping[event['EventSeverity']]
		c=ws[cellE_ID]
		if FormatWB:
			c.fill = PatternFill('solid', fgColor=rowfillcolor)
		ws[cellF_ID] = event['AppName']
		c=ws[cellF_ID]
		if FormatWB:
			c.fill = PatternFill('solid', fgColor=rowfillcolor)

		row = row + 1

